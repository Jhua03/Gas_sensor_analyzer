import matplotlib
matplotlib.use("TkAgg")  # Use TkAgg backend for Tkinter
import numpy as np
from matplotlib import pyplot as plt
import serial
import time
from matplotlib.widgets import TextBox, Button
from matplotlib.animation import FuncAnimation
from threading import Thread, Event
from openpyxl import Workbook
from scipy.signal import medfilt
from scipy.interpolate import interp1d
from scipy.optimize import curve_fit

# Variables
x_vals = []
y_vals = []
time_vals = []
currentState = Event()          
startTime = 0
input_neg_number = -1
input_pos_number = 1
input_step_number = 0.5
input_time_inverval_number = 1000
stop_time = 0
y_lim = [-1, 1]
x_lim = [-1, 1]
current_time = 0

#Workbook
#filename = "C:\SFU\CovidTestingLab\Findings\data2.xlsx"
#workbook = Workbook()
#worksheet = workbook.active
#row = 1
#Timecol = 1
#vDACcol = 2
#voutcol = 3
#worksheet.cell(row=row, column=Timecol, value="Time")
#worksheet.cell(row=row, column=vDACcol, value="VDAC")
#worksheet.cell(row=row, column=voutcol, value="Vout")
#workbook.save(filename=filename)
#row = 2
def on_button_clicked(event):
    global startTime, row,stop_time,ser
    #row = 2
    #worksheet.delete_rows(2, worksheet.max_row)
    
    if 'ser' not in globals() or not ser.is_open:
        try:
            ser = serial.Serial('COM3', 115200)
        except Exception as e:
            print(f"Error opening serial port: {e}")
            return
    init = "Start"
    ser.write(init.encode());
    startTime = time.time()
    currentState.set()           # Set the event to start data collection

def on_stop_button_clicked(event):
    currentState.clear()         # Clear the event to stop data collection
    ser.write("pause".encode())  # Send "pause" command to serial device
    
 
    
def on_clear_button_clicked(event):
    global x_vals, y_vals, time_vals
    ser.write("Clear".encode());
    if 'ser' in globals() and ser.is_open:
       ser.close()   # Close the serial port
    x_vals.clear()
    y_vals.clear()
    time_vals.clear()

    

def read_serial_data():
    global ser, startTime,row
    while not currentState.is_set():
        time.sleep(0.1)
    try:
        while True:
            if currentState.is_set():
                value = ser.readline().decode('utf-8').strip()
                dataArray = value.split(' ')
                x_vals.append(float(dataArray[1]))
                y_vals.append((float(dataArray[3]))/-100000)  
                current_time = time.time()
                time_vals.append(current_time - startTime)  
                #worksheet.cell(row=row, column=Timecol, value= float(current_time - startTime))
                #worksheet.cell(row=row, column=vDACcol, value= float(dataArray[1]))
                #worksheet.cell(row=row, column=voutcol, value= float(dataArray[3]))
                #workbook.save(filename=filename)
                #row = row + 1
                if len(x_vals) > 100 or len(y_vals) > 100 or len(time_vals) > 100:
                    x_vals.pop(0)
                    y_vals.pop(0)  
                    time_vals.pop(0)
                if len(x_vals) > 2 and (abs(input_pos_number - input_neg_number) -0.05) < (abs(x_vals[-1]-x_vals[0]) ) :
                    x_vals.pop(0)
                    y_vals.pop(0)  
                    time_vals.pop(0)
                    currentState.clear()
                    
                    
            else:
                time.sleep(0.01)
    except KeyboardInterrupt:
        pass
    except Exception as e:
        print(f"An error occurred: {e}")

    

def on_close(event):
    
    global x_vals, y_vals, time_vals
    x_vals.clear()
    y_vals.clear()
    time_vals.clear()
    currentState.clear()  # Stop the data collection
    if 'ser' in globals() and ser.is_open:
       ser.close()   # Close the serial port
    plt.close(fig) 

def on_log_button_clicked(event):
    global y_vals
    y_vals = [np.log10(y) if y > 0 else np.nan for y in y_vals]
    print("Converted y-values to logarithmic scale.")

def negative(text):
    global input_neg_number, ser
    try:
        if '.' not in text:
            text = text + '.'
        text = text.ljust(5, '0')[:5] 
        input_neg_number = float(text)
        print(f"Input number updated to: {text}")
        ser.write(text.encode())
    except ValueError:
        print("Invalid input. Please enter a valid number.")
        
def positive(text):
    global input_pos_number, ser
    try:
        if '.' not in text:
            text = text + '.'
        text = text.ljust(5, '0')[:5] 
        input_pos_number = float(text)
        print(f"Input number updated to: {text}")
        ser.write(text.encode())
    except ValueError:
        print("Invalid input. Please enter a valid number.")
        
def step(text):
    global input_step_number, ser
    try:
        text = 's' +text
        if '.' not in text:
            text = text + '.'
        text = text.ljust(5, '0')[:5] 
        ser.write(text.encode())
    except ValueError:
        print("Invalid input. Please enter a valid number.")
def time_interval(text):
    global input_time_interval_number, ser
    try:
        input_time_interval_number = float(text)
        text = 't' +text
        if '.' not in text:
            text = text + '.'
        text = text.ljust(5, '0')[:5] 
        print(f"Input number updated to: {text}")
        ser.write(text.encode())
    except ValueError:
        print("Invalid input. Please enter a valid number.")

def update_y_limits(text):
    global y_lim
    y_lim = list(map(float, text.split(',')))
    V_I.set_ylim(y_lim)
    plt.draw()

        
def update_x_limits(text):
    global x_lim
    x_lim = list(map(float, text.split(',')))
    V_I.set_xlim(x_lim)
    plt.draw()


# Create figure and axis
fig, (V_I, vDAC_time, vout_time) = plt.subplots(1, 3, figsize=(20, 5))
V_I.set_xlabel('vDAC')
V_I.set_ylabel('Current (nA)')
V_I.set_title('V-I Characteristic')
V_I.grid(which='both', color='gray', linestyle='dashed')
V_I.set_xlim(x_lim) 
V_I.set_ylim(y_lim)

vDAC_time.set_xlabel('Time')
vDAC_time.set_ylabel('vDAC')
vDAC_time.set_title('vDAC over Time')
vDAC_time.grid(which='both', color='gray', linestyle='dashed')
vDAC_time.set_xlim([0, 100]) 
vDAC_time.set_ylim(x_lim) 


vout_time.set_xlabel('Time')
vout_time.set_ylabel('Current (nA)')
vout_time.set_title('Current over Time')
vout_time.grid(which='both', color='gray', linestyle='dashed')
vout_time.set_ylim(y_lim) 

# Create start button
plt.subplots_adjust(bottom=0.2)
button_ax = plt.axes([0.3, 0.025, 0.2, 0.075])
button = Button(button_ax, 'Start')
button.on_clicked(on_button_clicked)

# Create clear button
clear_button_ax = plt.axes([0.7, 0.025, 0.2, 0.075])
clear_button = Button(clear_button_ax, 'Clear')
clear_button.on_clicked(on_clear_button_clicked)

# Create stop button
stop_button_ax = plt.axes([0.5, 0.025, 0.2, 0.075])
stop_button = Button(stop_button_ax, 'Stop')
stop_button.on_clicked(on_stop_button_clicked)

#Create negative box
negative_box_ax = plt.axes([0.025, 0.1, 0.05, 0.075])
negative_box = TextBox(negative_box_ax, 'Limit', initial=str(input_neg_number))
negative_box.on_submit(negative)

#Create positive box
positive_box_ax = plt.axes([0.075, 0.1, 0.05, 0.075])
positive_box = TextBox(positive_box_ax, '', initial=str(input_pos_number))
positive_box.on_submit(positive)

#Create step box
step_box_ax = plt.axes([0.025, 0.3, 0.05, 0.075])
step_box = TextBox(step_box_ax, 'Step', initial=str(input_step_number))
step_box.on_submit(step)

#Create time_interval box
time_interval_box_ax = plt.axes([0.025, 0.2, 0.05, 0.075])
time_interval_box = TextBox(time_interval_box_ax, 'Time', initial=str(input_time_inverval_number))
time_interval_box.on_submit(time_interval)

# Create log button
log_button_ax = plt.axes([0.91, 0.5, 0.05, 0.075])
log_button = Button(log_button_ax, 'Log')
log_button.on_clicked(on_log_button_clicked)


# Create TextBox for y limits
axbox_y = plt.axes([0.05, 0.8, 0.05, 0.075])
text_box_y = TextBox(axbox_y, 'Current Limits', initial=f"{y_lim[0]},{y_lim[1]}")
text_box_y.on_submit(update_y_limits)

# Create TextBox for x limits
axbox_x = plt.axes([0.05, 0.7, 0.05, 0.075])
text_box_x = TextBox(axbox_x, 'vDAC Limits', initial=f"{x_lim[0]},{x_lim[1]}")
text_box_x.on_submit(update_x_limits)

def linear_fit(x, a, b):
    return a * x + b


def moving_average(data, window_size):
    if len(data) < window_size:
        return data  
    return np.convolve(data, np.ones(window_size)/window_size, mode='valid')

def combined_filter(data, window_size, median_kernel_size):
    smoothed_data = moving_average(data, window_size=window_size)
    if len(smoothed_data) >= median_kernel_size:
        filtered_data = medfilt(smoothed_data, kernel_size=median_kernel_size)
    else:
        filtered_data = smoothed_data  
    return filtered_data

def remove_duplicates(x_vals, y_vals):
    """Remove duplicate x-values while keeping the first occurrence of each x-value."""
    unique_x, unique_indices = np.unique(x_vals, return_index=True)
    unique_y = [y_vals[i] for i in unique_indices]
    return unique_x, unique_y

def update_plot(frame):

    # Perform linear curve fitting
    x_fit = x_vals
    y_fit = y_vals
    popt = [1, 0] 
    if len(x_fit) > 3:  # Ensure enough data points for fitting
       try:
           popt, _ = curve_fit(linear_fit, x_fit, y_fit)
           x_fit_line = np.linspace(min(x_fit), max(x_fit), 500)
           y_fit_line = linear_fit(x_fit_line, *popt)
       except Exception as e:
           print(f"Curve fitting failed: {e}")
           x_fit_line, y_fit_line = x_fit, y_fit  # Fallback to raw data
    else:
       x_fit_line, y_fit_line = x_fit, y_fit  # Fallback to raw data
        
        
    V_I.clear()
    V_I.plot(x_vals,y_vals, label='Data')
    V_I.plot(x_fit_line, y_fit_line, '-', label=f'Fit: y = {popt[0]:.3f}x + {popt[1]:.3f}')
    V_I.set_xlabel('vDAC')
    V_I.set_ylabel('Current (nA)')
    V_I.set_title('V-I Characteristic')
    V_I.grid(which='both', color='gray', linestyle='dashed')
    #V_I.set_xlim(x_lim)
    V_I.legend()
   
    vDAC_time.clear()
    vDAC_time.plot(time_vals ,x_vals)
    vDAC_time.set_xlabel('Time')
    vDAC_time.set_ylabel('vDAC')
    vDAC_time.set_title('vDAC over Time')
    vDAC_time.grid(which='both', color='gray', linestyle='dashed')
    #vDAC_time.set_xlim([0, 100]) 
    #vDAC_time.set_ylim(x_lim) 

    vout_time.clear()
    vout_time.plot(time_vals, y_vals )
    vout_time.set_xlabel('Time')
    vout_time.set_ylabel('Current (nA)')
    vout_time.set_title('Current over Time')
    vout_time.grid(which='both', color='gray', linestyle='dashed')
    
    
    
# Start serial reading in a separate thread
serial_thread = Thread(target=read_serial_data)
serial_thread.daemon = True
serial_thread.start()

# Set up animation
ani = FuncAnimation(fig, update_plot, blit=False, interval=100)

# Connect the close event handler
fig.canvas.mpl_connect('close_event', on_close)

plt.show()

if 'ser' in globals() and ser is not None and ser.is_open:
    ser.close()
